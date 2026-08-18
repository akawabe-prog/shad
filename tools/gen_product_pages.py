# -*- coding: utf-8 -*-
"""
SHAD 商品詳細ページ ジェネレーター
- SHAD_ItemList.xlsx からモデル単位の詳細情報を抽出
- 共有アセットから「商品スチール優先」で画像を再選定（メイン+ギャラリー最大4）
- site/product-{code}.html を1モデルずつ生成
"""
import openpyxl, os, re, json, glob, subprocess, collections, html

BASE = '/Users/cjmac002/Desktop/RIDEOUT_MEDIA/SHAD_ReBranding'
A = f'{BASE}/assets'
SITE = f'{BASE}/site'
OUTIMG = f'{SITE}/img/products'
XLSX = '/Users/cjmac002/Downloads/SHAD_ItemList.xlsx'

MODELS = {  # code: (series, jp, copy, asset folder)
 'TR41':('TERRA','TERRA トップケース','硬化アルミ合金×統合ダブルロック。日常からアドベンチャーまで。','00 - NEW PRODUCTS/02_TR41'),
 'TR46':('TERRA','TERRA トップケース','TR41より15%大容量。ソフトトレイルの相棒。','00 - NEW PRODUCTS/04_TR46'),
 'TR48':('TERRA','TERRA トップケース','フルフェイス2個。TERRAの基準器。','01 - Products/1. Cases/TR48'),
 'TR55':('TERRA','TERRA トップケース','シリーズ最大55L。長旅のための容量。','01 - Products/1. Cases/TR55'),
 'TR37':('TERRA','TERRA トップケース','軽快な37L。街もトレイルも。','01 - Products/1. Cases/TR37'),
 'TR47':('TERRA','TERRA サイドケース','左右47L。重心を低く、積載を最大に。','01 - Products/1. Cases/TR47'),
 'TR36':('TERRA','TERRA サイドケース','スリムな36L。すり抜けを諦めない。','01 - Products/1. Cases/TR36'),
 'TR27':('TERRA','TERRA サイドケース','都市とオフロードのハイブリッド。','00 - NEW PRODUCTS/05-TR27'),
 'SH58X':('EXPANDABLE','可変容量トップケース','46L→58L。レバーひとつで3段階に伸縮。','01 - Products/1. Cases/SH58X & 59X - Social Media Images & Banners'),
 'SH38X':('EXPANDABLE','可変容量サイドケース','走るときはスリム。停まれば、大容量。世界でSHADだけ。','01 - Products/1. Cases/SH38X'),
 'SH48':('TOP CASE','トップケース','フルフェイス2個収納のベストセラー。',None),
 'SH51':('TOP CASE','トップケース','モジュラーヘルメット2個。ミドルアドベンチャーの最適解。','00 - NEW PRODUCTS/03_SH51'),
 'SH47':('TOP CASE','トップケース','エアロフォルムの47L。','01 - Products/1. Cases/SH47'),
 'SH44':('TOP CASE','トップケース','スマートロック搭載の44L。','01 - Products/1. Cases/SH44'),
 'SH33':('TOP CASE','トップケース','日常の定番33L。',None),
 'SH34':('TOP CASE','トップケース','スリム車体に馴染む34L。',None),
 'SH23':('SIDE CASE','サイドケース','車体に寄り添うスリムサイド。',None),
 'TR50':('BAG','TERRA リアバッグ','防水50L。ケースに代わる選択肢。','01 - Products/2. Bags/TR50'),
 'TR40':('BAG','アドベンチャーサイドバッグ','タフさとロックを兼ね備えたソフト。','01 - Products/2. Bags/TR40'),
 'TR30':('BAG','防水サイドバッグ','420D×両面PVC。豪雨でも中身は守る。','01 - Products/2. Bags/TR30'),
 'TR10':('BAG','クリックシステム タンクバッグ','ワンタッチ着脱。タンクに触れない。','01 - Products/2. Bags/03_TR10 - Click System'),
 'E48':('BAG','シートバッグ','クリックシステム対応の大容量。','01 - Products/2. Bags/E48'),
 'LOCK':('LOCK','SHADロック ハンドルロック','屈まない、汚れない。日常の防犯を数秒で。','01 - Products/4. SHAD Locks'),
 'SEAT':('SEAT','コンフォートシート','人間工学で疲労を設計から減らす。',None),
}
CAPS={'TR41':'41L','TR46':'46L','TR48':'48L','TR55':'55L','TR37':'37L','TR47':'47L','TR36':'36L','TR27':'27L',
 'SH58X':'46-58L','SH38X':'EXP','SH48':'48L','SH51':'51L','SH47':'47L','SH44':'44L','SH33':'33L','SH34':'34L',
 'SH23':'23L','TR50':'50L','TR40':'40L','TR30':'30L','TR10':'10L','E48':'48L','LOCK':None,'SEAT':None}
NEW={'TR46','TR27','SH51',}
FLAG={'SH38X','SH58X','TR41'}

# ヘルメットだけTablerに無いのでカスタムSVG（モノライン）
HELMET_SVG='<svg viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.6" stroke-linecap="round" stroke-linejoin="round" width="1em" height="1em"><path d="M3.5 14a8.5 8.5 0 0 1 17 0"/><path d="M3.5 14h17v1.5a1.5 1.5 0 0 1-1.5 1.5h-3.2l-.9 2.4a1 1 0 0 1-.94.6H10.0a1 1 0 0 1-.94-.6L8.2 17H5a1.5 1.5 0 0 1-1.5-1.5z"/><path d="M9 14c0-3 1-5 4.5-5"/></svg>'

# 各モデルの特徴ピクトグラム（本国サイト流・瞬時にわかる）
# 形式: (icon, ラベル, 値 or None)  icon='helmet' はカスタム、他は Tabler の ti-{icon}
FEAT={
 'TR41':[('box','容量','41L'),('helmet','ヘルメット','×2'),('lock','統合ダブルロック',None),('shield','アルミ',None)],
 'TR46':[('box','容量','46L'),('helmet','ヘルメット','×2'),('lock','統合ダブルロック',None),('shield','ポリプロピレン',None)],
 'TR48':[('box','容量','48L'),('helmet','ヘルメット','×2'),('lock','統合ダブルロック',None),('shield','アルミ',None)],
 'TR55':[('box','容量','55L'),('helmet','ヘルメット','×2'),('lock','統合ダブルロック',None),('shield','アルミ',None)],
 'TR37':[('box','容量','37L'),('helmet','ヘルメット','×1'),('lock','統合ダブルロック',None),('shield','アルミ',None)],
 'TR47':[('box','容量','47L'),('tool','3P/4Pマウント',None),('lock','統合ダブルロック',None),('shield','アルミ',None)],
 'TR36':[('box','容量','36L'),('tool','3P/4Pマウント',None),('lock','統合ダブルロック',None),('shield','アルミ',None)],
 'TR27':[('box','容量','27L'),('tool','3P/4Pマウント',None),('lock','統合ダブルロック',None)],
 'SH58X':[('arrows-maximize','可変容量','46–58L'),('key','スマートロック',None),('helmet','ヘルメット','×2'),('shield','カーボン/アルミ',None)],
 'SH38X':[('arrows-maximize','可変容量',None),('tool','3P/4Pマウント',None),('key','スマートロック',None)],
 'SH48':[('box','容量','48L'),('helmet','ヘルメット','×2'),('key','キーレス開閉',None),('weight','耐荷重','8kg')],
 'SH51':[('box','容量','51L'),('helmet','ヘルメット','×2'),('key','スマートロック',None),('weight','耐荷重','10kg')],
 'SH47':[('box','容量','47L'),('helmet','ヘルメット','×2'),('key','スマートロック',None)],
 'SH44':[('box','容量','44L'),('helmet','ヘルメット','×1'),('key','スマートロック',None)],
 'SH33':[('box','容量','33L'),('helmet','ヘルメット','×1'),('key','スマートロック',None)],
 'SH34':[('box','容量','34L'),('helmet','ヘルメット','×1'),('key','スマートロック',None)],
 'SH23':[('box','容量','23L'),('tool','3P/4Pマウント',None),('shield','スリム設計',None)],
 'TR50':[('box','容量','50L'),('droplet','防水',None),('lock','ロック機能',None)],
 'TR40':[('box','容量','40L'),('droplet','防水',None),('lock','ロック機能',None),('sbh','サイドバッグホルダー',None)],
 'TR30':[('box','容量','60L'),('droplet','防水','IPX6'),('weight','耐荷重','6kg'),('sbh','サイドバッグホルダー',None)],
 'TR10':[('box','容量','10L'),('click','クリックシステム',None),('droplet','防水',None)],
 'E48':[('box','容量','48L'),('click','クリックシステム',None),('shield','シートバッグ',None)],
 'LOCK':[('key','キーで施錠',None),('click','工具不要',None),('shield-lock','防犯',None)],
 'SEAT':[('armchair','人間工学',None),('shield','振動吸収',None),('shield-check','純正フィット',None)],
}
# 本国公式ディスクリプションの日本語版（Excelの汎用説明を上書き）
DESC={
 'SH38X':'エンジニアリングとデザインが融合したSH38X Expandable。その核心は、建築の概念「テンセグリティ」に着想を得た可変容量メカニズムにあります。\n'
         'この技術をモーターサイクルの要求に最適化し、これまでにない新概念「ダイナミック・テンセグリティ」を実現しました。張力を相互に補償する革新的な機械システムにより、ケースはわずか数秒で、均一に、テレスコピック（伸縮）展開します。その先進的なデザインと技術は、栄えあるRed Dot デザイン賞 2024を受賞しました。\n'
         'SH38Xは+70mm伸長し、容積が40%拡大。XLサイズのモジュラーヘルメット（ECE 22-06）を収納できます。これにより、ヘルメットが収まる可変容量サイドケースとして、市場で最もコンパクトな一台となりました。',
 'SH58X':'市場唯一の可変容量トップケース、SH58X。革新的な設計で荷室容量を数秒のうちに3段階（46・52・58L）へ調整できます。気密ロック「スマートロックシステム」が最高レベルの防犯性を実現。カバーのカラー展開で、アスファルトもスポーツツーリングも自在に演出できます。',
 'SH48':'市場の要求をすべて満たすハイエンド・トップケース。わずか3.7kgで、48Lの大容量クラスでは市場最軽量です。軽さが振動を抑え、ステーと車体フレームへの負担を軽減。「スマートロックシステム」で高い防犯性を備え、多彩なカスタマイズ構成にも対応します。',
 'SH51':'アドベンチャー・ツーリングのために設計されたトップケース、SH51。高耐久素材と空力デザインが走行時の安全性と安定性を提供します。フリップアップ2個を収納でき、小物整理用の内装メッシュを完備。最大積載10kg。SH38X Expandableサイドケースと組み合わせれば、大容量が必要なロングツーリングに最適なセットになり、コンパクトな構成への変更も容易です。',
 'SH47':'バルセロナで設計されたSH47。アグレッシブな空力フォルムに、モダンで洗練されたルックスを与えました。街乗りからロングライドまで、スクーターから中・大型車まで幅広く適合する万能サイズです。',
 'SH44':'フルフェイス2個を収納できるトップケース。バイクにもスクーターにも最適です。',
 'SH33':'フルフェイス1個と小物が収まる33Lトップケース。コンパクトで実用的、豊富なカラー展開も魅力です。軽く押すだけで閉まる「プレスロックシステム」を搭載しています。',
 'SH34':'フルフェイス1個と小物が収まる34Lトップケース。コンパクトで日常使いに最適。「プレスロックシステム」を搭載しています。',
 'SH23':'エレガンス・軽量・機能性を兼ね備えたスリムなサイドケース、SH23。優れた空力性能を保ちながら、耐久性と防水性も確保。23Lの容量をすべて使い切れ、上開き構造でバイクを降りずに荷物へアクセスできます。',
 'TR41':'アドベンチャーのために設計されたタフなケース、TR41。高強度素材と、上部フレーム＋中央ボディ構造で補強した堅牢な基部設計が特長です。TR27サイドケースとの組み合わせに最適。ステンレス製ロックと一体型ハンドル、アルミカバーを備えます。JETヘルメット2個を収納、最大積載5kg。堅牢な構造と機能美が評価され、Red Dot デザイン賞 2026を受賞しました。',
 'TR46':'アドベンチャー向けのタフなケース、TR46。高強度素材と補強された基部設計で堅牢性を実現。TR27サイドケースとの相性も抜群です。ステンレス製ロックと一体型ハンドルを備え、フルフェイス＋JETの2個を収納、最大積載6kg。Red Dot デザイン賞 2026を受賞しています。',
 'TR48':'設計・製造を100%バルセロナで行うTR48トップケース。トレイル用ヘルメット1個、またはフリップアップ＋ジェットの2個を収納します。硬化アルミ合金製で、強く軽い素材により1.2mmの薄肉断面を実現し、最小重量で最大の耐久性を両立。特許「TERRAロックシステム」を搭載します。※取付プレートは別売。',
 'TR55':'フリップアップ2個を収納するTERRA TR55トップケース。5.4kgとアルミケース市場最軽量で、設計・製造は100%バルセロナ。革新的デザインで車体への視覚的インパクトを最小化しました。TERRAを基準たらしめた特許「TERRAロックシステム」など全技術を搭載し、デザイン性でABCアワードを受賞しています。※取付プレートは別売。',
 'TR27':'アドベンチャー走行のために設計されたサイドケース、TR27。高耐久素材とリブ構造による補強で、TR41・TR46・SH51トップケースと完璧にマッチします。3P/4Pシステムに対応し、上部固定ポイントで積載量を拡張可能、片側最大7kg。アスファルトもオフロードも走破する汎用性が最大の魅力です。',
 'TR50':'フルフェイス2個（40L）・最大10kgを積めるTERRA TR50リアバッグ。ロック付きフィッティングで数秒で固定・着脱できるのが特長です。※取付プレートは別売。',
 'TR40':'バルセロナで設計された容量64Lのアドベンチャーサイドバッグ、TR40。高耐摩耗素材を採用し、ロール式クロージャーとインナーバッグの組み合わせで防水を確保。特許「ダブルロックシステム」が、バッグのフィッティングへの固定とバッグ自体の施錠を両立し、内部への侵入を防ぎます。',
 'TR30':'完全防水（IPX6）のアドベンチャーサイドバッグ、TR30。バルセロナ設計で容量は最大60L。特許「ダブルロックシステム」がバッグを4P/3Pフィッティングへ固定すると同時に、バッグ自体も施錠して内部への侵入を防ぎます。',
 'E48':'容量46〜58Lの可変式サイドバッグ（2個）。フルフェイス1個を収納できます。SHAD SR・SBHフィッティング（適合車種）、または汎用ベルトの3通りで車体に装着可能です。',
}

# LP特徴ストーリー（全モデル分は story_data.py に分離）
import sys as _sys; _sys.path.insert(0, os.path.dirname(os.path.abspath(__file__)))
from story_data import STORY
_UNUSED_STORY={
 'SH38X':[
   {'k':'Expandable','t':'走るときはスリム。停まれば、大容量。','img':'img/products/sh38x_story1.jpg',
    'body':'レバー操作で数秒。走行時は車幅を抑えてスリムに、停車後はフルフェイスが収まる容量へ。容量が変わるサイドケースは、世界でSHADだけ。'},
   {'k':'3P / 4P System','t':'車体に、溶け込むマウント。','img':'img/products/sh38x_story2.jpg',
    'body':'必要なアンカーを1本のアームで支える特許構造。マスを車体中心に集め、走行中のハンドリングへの影響を最小化。ケースを外しても美しいシルエットのまま。'},
   {'k':'Premium Smart Lock','t':'鍵いらずの、ワンタッチ。','img':'img/products/sh38x_story3.jpg',
    'body':'解錠状態にしておけば、ボタン操作だけで開閉。料金所での小銭、休憩時の出し入れ。片手で完結する操作性が、毎日の所作を変える。'},
 ],
 'TR41':[
   {'k':'Hardened Aluminium','t':'アルミの堅牢を、旅の標準に。','img':'img/products/tr41_story1.jpg',
    'body':'硬化アルミ合金と、上部フレーム＋中央ボディで補強した基部構造。過酷なアドベンチャーに耐える堅牢さが、その機能美とともにRed Dot デザイン賞 2026に評価された。'},
   {'k':'TERRA Lock System','t':'鍵ひとつで、開閉も着脱も。','img':'img/products/tr41_story2.jpg',
    'body':'ステンレス製ロックと一体型ハンドルを備えた特許ロックシステム。手袋のままでも確実に扱え、車体からの取り外しも素早く。'},
   {'k':'Adventure Ready','t':'TR27と組めば、旅仕様へ。','img':'img/products/tr41_story3.jpg',
    'body':'JETヘルメット2個を収納、最大積載5kg。TR27サイドケースと組み合わせれば、長距離アドベンチャーに必要な積載を一式で揃えられる。'},
 ],
 'TR46':[
   {'k':'Built Tough','t':'アドベンチャーのための、タフネス。','img':'img/products/tr46_story1.jpg',
    'body':'高強度素材と補強された基部設計で、堅牢性を徹底追求。TR41譲りの構造に、扱いやすさを両立した。機能美はRed Dot デザイン賞 2026に評価されている。'},
   {'k':'2 Helmets','t':'フルフェイスとジェット、2個入る。','img':'img/products/tr46_story2.jpg',
    'body':'フルフェイス＋ジェットの2個を収納でき、最大積載6kg。TR41より15%大きな容量で、日常から旅まで受け止める。'},
   {'k':'TERRA Lock System','t':'確実に、素早く。','img':'img/products/tr46_story3.jpg',
    'body':'ステンレス製ロックと一体型ハンドル。開閉も車体への着脱も、片手で確実に行える。TR27サイドケースとの組み合わせにも最適。'},
 ],
}
def feat_icon(k, cls=''):
    if k=='helmet': return f'<span class="{cls}" aria-hidden="true">{HELMET_SVG}</span>'
    k={'sbh':'luggage'}.get(k,k)  # 公式画像が無い場合のモノライン代替
    return f'<i class="ti ti-{k} {cls}" aria-hidden="true"></i>'

# 本国公式アイコン（site/img/feat/*.png＝透過インク加工済み）へのマッピング
def official_for(ic, val, label):
    name=None
    if ic=='helmet':            name='2CI' if (val and '2' in str(val)) else '1CI'
    elif ic=='arrows-maximize': name='EX'
    elif ic=='click':           name='CS'
    elif ic=='tool':            name='4P' if '4P' in (label or '') and not '3P' in (label or '') else '3P'
    elif ic=='sbh':             name='SSBHE'
    elif ic=='weight' and val:
        n=re.sub(r'[^0-9]','',str(val)); name='MaxLoad'+n if n else None
    if name and os.path.exists(f'{SITE}/img/feat/{name}.png'):
        return f'img/feat/{name}.png'
    return ''

def feat_node_html(f, big=False):
    """公式アイコンがあれば画像、無ければモノライン"""
    if f.get('oimg'):
        return f'<img src="{f["oimg"]}" alt="" class="feat-oimg" loading="lazy">'
    return feat_icon(f['ic'])

# ---------- Excel ----------
wb=openpyxl.load_workbook(XLSX, data_only=True)
ws=wb['Sheet1']; hdr=[c.value for c in ws[1]]; ix={h:i for i,h in enumerate(hdr)}
MODRE=re.compile(r'\b(SH\d{2}X?|TR\d{2}|SW80|SL\d{2}|E\d{2}(?:SR)?)\b', re.I)

def is_excluded(g):
    """SHAD商品リスト 正式除外ルール（2026-06-15ユーザー指定）"""
    if str(g('CJ廃番'))=='1': return True                       # ① CJ廃番=1
    if '【セット品】' in str(g('商品名') or ''): return True      # ② セット品
    hb=str(g('品番') or '').upper()
    if hb.startswith('YY') or hb.startswith('ZZ'): return True   # ③ 品番 YY/ZZ始まり
    return False

rows_by_model=collections.defaultdict(list)
for r in ws.iter_rows(min_row=2, values_only=True):
    g=lambda k: r[ix[k]]
    if is_excluded(g): continue
    name=str(g('商品名') or ''); mpn=str(g('メーカー品番') or '')
    cat=str(g('カテゴリ名') or '')
    m=MODRE.search(mpn) or MODRE.search(name)
    code=None
    if m: code=m.group(1).upper()
    if cat=='ハンドルロック': code='LOCK'
    if cat=='シート': code='SEAT'
    if code not in MODELS: continue
    rows_by_model[code].append({k: (str(g(k)) if g(k) is not None else '') for k in
        ('商品名','メーカー品番','JANコード','メインカラー','キャッチ','商品説明メイン','商品説明サブ',
         '容量','質量','材質','商品サイズ','仕様','セット内容・付属品','注意','カテゴリ名')})

# ---------- 画像選定（商品スチール優先） ----------
jan_dir=f'{A}/Website_material_CSV_&_images/Images'
jan_files=collections.defaultdict(list)
for fn in sorted(os.listdir(jan_dir)):
    m=re.match(r'(\d+)_(\d+)\.jpg', fn)
    if m: jan_files[m.group(1)].append(fn)

# 除外：販促/動画/書類/ロゴ/各種バナーサイズ/ディテール接写
EXCLUDE=re.compile(r'(?i)social[ _]media|/banners?/|banner|/ads?/|video|product.?sheet|user.?guide|instruction|/texts?/|\.png$|logo|reel|1980|1920x|1080x|728x?|300x|250px|160x|320x|capacity|detalle|accesorio|_net|_key')
# ライフスタイル/装着写真（白背景でない）を完全除外
LIFESTYLE=re.compile(r'(?i)AC_|motorbike|africa.?twin|crf|_gs_|gs.?1250|gs.?1300|z650|trk|xadv|_jr_|moto_|ambient|amb_|scene|ride|lifestyle|action|kawasaki|bmw|honda|yamaha|suzuki|ducati|benelli|triumph')

# 本体SKUのJAN（assets全体に _1.. 画像があるもの）を事前解決
JAN_IDX=collections.defaultdict(list)
for _p in glob.glob(f'{A}/**/*.jpg', recursive=True):
    _m=re.match(r'(\d{13})_(\d+)\.jpg$', os.path.basename(_p))
    if _m: JAN_IDX[_m.group(1)].append((int(_m.group(2)), _p))
ACC_WORDS=re.compile(r'(?i)インナー|バックレスト|ストッパー|シリンダー|ステッカー|メッシュ|ダッフル|キー|アダプタ|ボルト|ロックセット|フィッティング|ベース|ステー|専用')
def body_jans(code):
    out=[]
    for row in rows_by_model.get(code,[]):
        nm=row['商品名']
        if ACC_WORDS.search(nm): continue
        if code not in ('LOCK','SEAT') and code not in nm.upper().replace(' ',''): continue
        j=row['JANコード']
        if j in JAN_IDX: out.append(j)
    return out

def token_re(code):
    # SH38X / TR48 などをファイル名内で厳密一致（区切り許容）
    return re.compile(r'(?<![A-Z0-9])'+re.escape(code)+r'(?![A-Z0-9])', re.I)

# 自動解決できない本体（非表示/廃番SKU）の明示オーバーライド（assets相対のglob）
OVERRIDE={
 'SH58X':'_derived/sh58x_*.jpg',   # フラッグシップ：microsite画像からSH58X/59Xを切り出し
 'TR40':'01 - Products/2. Bags/TR40/Images/8430358680531_*.jpg',
 'E48' :'01 - Products/2. Bags/E48/Images/Images Product/8430358680944_*.jpg',
}
def _num(fn):
    m=re.search(r'(\d+)(?=\.jpg$)', fn); return int(m.group(1)) if m else 999

def pick_images(code):
    if code in OVERRIDE:
        ps=sorted(glob.glob(f'{A}/{OVERRIDE[code]}'), key=lambda p:_num(os.path.basename(p)))
        if ps: return ps[:4]
    cand=[]
    # 1) 公式EC商品画像（{JAN}_n）— 白背景・本体SKU。最優先
    for j in body_jans(code):
        for n,p in sorted(JAN_IDX[j]):
            cand.append((2000-n, p))
    # 2) ファイル名にモデル名を含む白背景スタジオ写真（assets全体・ライフスタイル除外）
    tok=token_re(code)
    for p in glob.glob(f'{A}/**/*.jpg', recursive=True):
        fn=os.path.basename(p)
        rel=os.path.relpath(p, A)   # 絶対パス（RIDEOUT等）の誤マッチ回避
        if EXCLUDE.search(rel) or LIFESTYLE.search(rel): continue
        if not tok.search(fn): continue
        s=1000
        m=re.match(r'\s*(\d{1,2})[\s_.-]', fn)
        if m: s-=int(m.group(1))                         # 連番01,02..の順
        if re.search(r'(?i)general|exterior|4000', fn): s+=20
        cand.append((s, p))
    seen=set(); out=[]
    for sc,p in sorted(cand, key=lambda x:-x[0]):
        key=os.path.basename(p)
        if key in seen: continue
        seen.add(key); out.append(p)
        if len(out)>=4: break
    return out

# ---------- 詳細データ構築 ----------
def rep_row(code):
    rs=rows_by_model.get(code,[])
    rs=[r for r in rs if r['カテゴリ名'] not in ('フィッティングキット・ステー・ベース',)]
    if not rs: return {}
    rs.sort(key=lambda r:(-len(r['商品説明メイン']), -len(r['仕様'])))
    return rs[0]

def colors_of(code):
    cols=[]
    for r in rows_by_model.get(code,[]):
        if r['カテゴリ名']=='フィッティングキット・ステー・ベース': continue
        c=r['メインカラー']
        if c and c not in cols: cols.append(c)
    return cols

def br(t):
    t=html.escape(t).replace('\\n','\n')
    return '<br>'.join([x for x in t.split('\n') if x.strip()])

products=[]
for code,(series,jp,copy_,folder) in MODELS.items():
    imgs=pick_images(code)
    if not imgs:
        print('SKIP', code); continue
    # 本国公式の看板画像があれば先頭に採用（白背景・高解像度・通販向け）
    official=f'{SITE}/img/official/{code.lower()}.jpg'
    if os.path.exists(official):
        imgs=[official]+[x for x in imgs if os.path.abspath(x)!=os.path.abspath(official)]
    paths=[]
    for i,src in enumerate(imgs[:4]):
        dst=f'{OUTIMG}/{code.lower()}_{i}.jpg'
        subprocess.run(['cp',src,dst],check=True)
        subprocess.run(['sips','-Z','1100' if i==0 else '760',dst],check=True,capture_output=True)
        paths.append(f'img/products/{code.lower()}_{i}.jpg')
    # 一覧カード用メイン
    subprocess.run(['cp',imgs[0],f'{OUTIMG}/{code.lower()}.jpg'],check=True)
    subprocess.run(['sips','-Z','900',f'{OUTIMG}/{code.lower()}.jpg'],check=True,capture_output=True)
    rr=rep_row(code)
    products.append({
      'code':code,'series':series,'jp':jp,'cap':CAPS[code],'copy':copy_,
      'colors':len(colors_of(code)),'img':f'img/products/{code.lower()}.jpg',
      'new':code in NEW,'flag':code in FLAG,
      'gallery':paths,'colorNames':colors_of(code),
      'desc':DESC.get(code) or rr.get('商品説明メイン',''),
      'desc2':'' if code in DESC else rr.get('商品説明サブ',''),
      'spec':{k:rr.get(k,'') for k in ('容量','質量','材質','商品サイズ','仕様','セット内容・付属品')},
      'note':rr.get('注意',''),
      'features':[{'ic':f[0],'label':f[1],'val':f[2],'oimg':official_for(f[0],f[2],f[1])} for f in FEAT.get(code,[])],
    })
    print('OK', code, len(paths), 'imgs |', os.path.basename(imgs[0])[:40])

# 一覧用JSON（軽量フィールドのみ）
lite=[{k:p[k] for k in ('code','series','jp','cap','copy','colors','img','new','flag','features')} for p in products]
json.dump(lite, open(f'{SITE}/products_data.json','w',encoding='utf-8'), ensure_ascii=False, indent=1)

# ---------- 詳細ページ生成 ----------
NAV='''<nav class="sticky top-0 z-50 bg-[rgba(10,10,10,.92)] backdrop-blur-md text-white">
  <div class="max-w-site mx-auto px-7 h-[72px] flex items-center gap-10">
    <a href="index.html" class="shrink-0"><img src="img/logo_slogan.svg" alt="SHAD" class="h-10 w-auto"></a>
    <ul class="hidden md:flex gap-8 flex-1 font-disp uppercase tracking-[.14em] text-[15.5px] font-medium text-white/80">
      <li><a class="text-white" href="products.html">Products</a></li>
      <li><a class="hover:text-white transition" href="index.html#finder">For Your Motorcycle</a></li>
      <li><a class="hover:text-white transition" href="index.html#why">Technology</a></li>
      <li><a class="hover:text-white transition" href="index.html#story">Brand</a></li>
      <li><a class="hover:text-white transition" href="index.html#store">Store</a></li>
    </ul>
    <div class="flex items-center gap-6 ml-auto md:ml-0">
      <a href="#" aria-label="カート" class="relative text-white/80 hover:text-white transition"><svg xmlns="http://www.w3.org/2000/svg" width="22" height="22" viewBox="0 0 24 24" fill="none" stroke="currentColor" stroke-width="1.8" stroke-linecap="round" stroke-linejoin="round" aria-hidden="true"><path d="M6 19m-2 0a2 2 0 1 0 4 0a2 2 0 1 0 -4 0"/><path d="M17 19m-2 0a2 2 0 1 0 4 0a2 2 0 1 0 -4 0"/><path d="M17 17h-11v-14h-2"/><path d="M6 5l14 1l-1 7h-13"/></svg><span class="absolute -top-1.5 -right-2 w-[15px] h-[15px] bg-shad rounded-full text-[9.5px] leading-[15px] text-center font-bold text-white">0</span></a>
    </div>
  </div>
</nav>'''

FOOTER='''<footer class="bg-[#0A0A0A] text-white pt-14 pb-8 mt-20">
  <div class="max-w-site mx-auto px-7">
    <img src="img/logo_slogan.svg" alt="SHAD — Engineered for Riding" class="h-[48px] w-auto">
    <div class="border-t border-white/10 mt-8 pt-6 flex items-center gap-4 text-[12.5px] text-white/55">
      <a href="products.html" class="hover:text-white transition">製品一覧へ戻る</a>
      <span class="ml-auto">© SHAD JAPAN — 日本総代理店 株式会社カスタムジャパン</span>
    </div>
  </div>
</footer>'''

def spec_rows(spec):
    LBL={'容量':'容量','質量':'質量','材質':'材質','商品サイズ':'サイズ','仕様':'仕様','セット内容・付属品':'セット内容・付属品'}
    out=''
    for k,l in LBL.items():
        v=spec.get(k,'')
        if not v.strip(): continue
        out+=f'<tr class="border-b border-black/10"><th class="text-left align-top py-3 pr-6 font-medium text-neutral-500 w-[140px] whitespace-nowrap">{l}</th><td class="py-3 text-[14px] leading-relaxed">{br(v)}</td></tr>'
    return out

for p in products:
    code=p['code']
    cap_html=''
    if p['cap']=='EXP':
        cap_html='<span class="cap-num" style="font-size:30px;">EXPANDABLE</span>'
    elif p['cap']:
        m=re.match(r'^([\d-]+)(L)$',p['cap'])
        cap_html=f'<span class="cap-num" style="font-size:56px;">{m.group(1)}<small style="font-size:28px;">L</small></span>' if m else f'<span class="cap-num">{p["cap"]}</span>'
    thumbs=''.join(f'<button class="g-thumb{" on" if i==0 else ""}" data-src="{src}"><img src="{src}" alt=""></button>' for i,src in enumerate(p['gallery']))
    colors=''.join(f'<span class="inline-block border border-black/20 rounded-full px-4 py-1.5 text-[13px]">{html.escape(c)}</span>' for c in p['colorNames'])
    badges=('<span class="new-badge" style="position:static;display:inline-block;">New</span> ' if p['new'] else '')+('<span class="feat-tag" style="background:#0E0E0E;color:#fff;display:inline-block;">Flagship</span>' if p['flag'] else '')
    related=[q for q in products if q['series']==p['series'] and q['code']!=code][:3]
    rel_html=''.join(f'''<a href="product-{q['code'].lower()}.html" class="pcard group bg-white rounded-[14px] overflow-hidden border border-black/10 transition hover:-translate-y-1 hover:shadow-[0_18px_40px_rgba(0,0,0,.10)]">
      <span class="block aspect-square overflow-hidden bg-white"><img src="{q['img']}" alt="{q['code']}" class="w-full h-full object-cover transition duration-300 group-hover:scale-[1.04]"></span>
      <span class="block px-5 py-4"><span class="font-disp font-semibold text-[20px] tracking-[.05em] uppercase">{q['code']}</span>
      <span class="block text-[12.5px] text-neutral-500 mt-0.5">{q['jp']}</span></span></a>''' for q in related)
    desc=br(p['desc']) or html.escape(p['copy'])
    desc2=br(p['desc2'])
    note=br(p['note'])

    # ===== LP型セクション（素材があるモデルのみ展開） =====
    lc=code.lower()
    mdir=f'{SITE}/media/products/{lc}'
    # ① ヒーロー動画
    lp_hero=''
    if os.path.exists(f'{mdir}/hero.mp4'):
        lp_hero=f'''<header class="lp-hero">
  <video src="media/products/{lc}/hero.mp4" autoplay muted loop playsinline></video>
  <div class="lp-hero-shade"></div>
  <div class="lp-hero-in">
    <p class="lp-kick">{p['series']}</p>
    <h1 class="lp-h1">{code}</h1>
    <p class="lp-hsub">{html.escape(p['copy'])}</p>
  </div>
  <span class="lp-scroll" aria-hidden="true"><i class="ti ti-chevron-down"></i></span>
</header>'''
    # ② 特徴ストーリー（画像×コピーの交互ブロック）
    lp_story=''
    if STORY.get(code):
        gal=p['gallery']
        # 画像未指定ブロックはギャラリー（0=看板を除く 1..）から循環割当
        pool=gal[1:] if len(gal)>1 else gal
        blocks=''
        for i,b in enumerate(STORY[code]):
            rev=' lp-rev' if i%2 else ''
            bimg=b.get('img') or (pool[i % len(pool)] if pool else gal[0])
            blocks+=f'''<div class="lp-block{rev}">
      <div class="lp-block-img"><img src="{bimg}" alt="{html.escape(b['t'])}" loading="lazy"></div>
      <div class="lp-block-tx"><span class="lp-block-kick">{html.escape(b.get('k',''))}</span>
        <h3 class="lp-block-h">{html.escape(b['t'])}</h3>
        <p class="lp-block-p">{html.escape(b['body'])}</p></div>
    </div>'''
        lp_story=f'<section class="lp-story"><div class="max-w-site mx-auto px-7">{blocks}</div></section>'
    # ③ 縦リール（SNS素材）
    reels=sorted(glob.glob(f'{mdir}/reel*.mp4'))
    lp_reels=''
    if reels:
        tiles=''.join(f'<div class="lp-reel"><video src="media/products/{lc}/{os.path.basename(r)}" muted loop playsinline preload="none"></video></div>' for r in reels)
        lp_reels=f'''<section class="lp-reels"><div class="max-w-site mx-auto px-7">
      <h2 class="sec-ttl sec-ttl-quiet text-white">On the Road</h2>
      <div class="lp-reel-row">{tiles}</div></div></section>'''
    # ④ サポート（ユーザーガイドDL）
    lp_support=''
    if os.path.exists(f'{SITE}/docs/{lc}_userguide.pdf'):
        lp_support=f'''<section class="max-w-site mx-auto px-7 py-12">
      <div class="lp-support">
        <div><span class="font-disp text-[13px] tracking-[.22em] uppercase text-neutral-400">Support</span>
          <h2 class="text-[22px] font-bold mt-1">取り付けは、説明書とプロにお任せ。</h2>
          <p class="text-[13.5px] text-neutral-500 mt-2">公式ユーザーガイドで取付手順を確認できます。装着は適合確認のうえ、最寄りの取扱店でも承ります。</p></div>
        <div class="lp-support-act">
          <a href="docs/{lc}_userguide.pdf" target="_blank" class="btn bg-ink text-white hover:bg-black !py-3 !px-6 !text-[14px]"><i class="ti ti-file-type-pdf"></i>ユーザーガイド（PDF）</a>
          <a href="index.html#store" class="btn border border-black/25 text-ink hover:bg-mist !py-3 !px-6 !text-[14px]">取扱店を探す</a>
        </div>
      </div></section>'''
    # 特徴ピクトグラム（詳細ページ・大きめ）
    feat_html=''
    if p['features']:
        cells=''.join(
          f'<div class="feat-cell"><span class="feat-ic">{feat_node_html(f)}</span>'
          f'<span class="feat-tx"><b>{html.escape(f["label"])}</b>'
          + (f'<span>{html.escape(f["val"])}</span>' if f.get("val") else '')
          + '</span></div>'
          for f in p['features'])
        feat_html=f'<div class="feat-grid">{cells}</div>'
    page=f'''<!doctype html>
<html lang="ja">
<head>
<meta charset="utf-8"><meta name="viewport" content="width=device-width, initial-scale=1">
<title>{code}｜{p['jp']} — SHAD JAPAN</title>
<meta name="description" content="{html.escape(p['copy'])}">
<link rel="preconnect" href="https://fonts.googleapis.com"><link rel="preconnect" href="https://fonts.gstatic.com" crossorigin>
<link href="https://fonts.googleapis.com/css2?family=Barlow+Condensed:wght@500;600&family=Noto+Sans+JP:wght@400;500;700&display=swap" rel="stylesheet">
<link rel="stylesheet" href="https://cdn.jsdelivr.net/npm/@tabler/icons-webfont@3.24.0/dist/tabler-icons.min.css">
<script src="https://cdn.tailwindcss.com"></script>
<script>tailwind.config={{theme:{{extend:{{colors:{{shad:'#E31E24',ink:'#0E0E0E',ink2:'#161616',mist:'#F4F3F1'}},fontFamily:{{disp:['"Barlow Condensed"','"Noto Sans JP"','sans-serif'],sans:['"Noto Sans JP"','"Hiragino Kaku Gothic ProN"','sans-serif']}},maxWidth:{{site:'1180px'}}}}}}}};</script>
<link rel="stylesheet" href="css/custom.css">
<style>
.cap-num{{font-family:"Barlow Condensed",sans-serif;font-weight:600;line-height:1;letter-spacing:.02em;color:#E31E24;}}
.g-main{{aspect-ratio:1/1;background:#fff;border:1px solid rgba(0,0,0,.1);border-radius:14px;overflow:hidden;}}
.g-main img{{width:100%;height:100%;object-fit:cover;}}
.g-thumb{{width:72px;height:72px;border-radius:10px;overflow:hidden;border:2px solid rgba(0,0,0,.12);background:#fff;cursor:pointer;transition:.15s;}}
.g-thumb img{{width:100%;height:100%;object-fit:cover;}}
.g-thumb.on{{border-color:#E31E24;}}
.new-badge{{background:#E31E24;color:#fff;font-family:"Barlow Condensed",sans-serif;font-weight:600;font-size:11.5px;letter-spacing:.18em;text-transform:uppercase;padding:3px 9px;border-radius:4px;}}
.feat-tag{{font-family:"Barlow Condensed",sans-serif;font-weight:600;font-size:11px;letter-spacing:.2em;text-transform:uppercase;padding:2.5px 8px;border-radius:4px;}}
.feat-grid{{display:grid;grid-template-columns:repeat(2,1fr);gap:14px;margin-top:22px;border-top:1px solid rgba(0,0,0,.1);padding-top:22px;}}
.feat-cell{{display:flex;align-items:center;gap:12px;}}
.feat-ic{{flex:none;width:42px;height:42px;border-radius:10px;background:var(--mist,#F4F3F1);display:flex;align-items:center;justify-content:center;font-size:22px;color:#0E0E0E;}}
.feat-ic svg{{width:22px;height:22px;}}
.feat-oimg{{width:30px;height:30px;object-fit:contain;}}
.feat-tx{{display:flex;flex-direction:column;line-height:1.35;}}
.feat-tx b{{font-weight:500;font-size:13.5px;}}
.feat-tx span{{font-family:"Barlow Condensed",sans-serif;font-weight:600;font-size:17px;color:#E31E24;letter-spacing:.02em;}}
/* LP hero */
.lp-hero{{position:relative;background:#000;color:#fff;aspect-ratio:16/7;width:100%;display:flex;align-items:flex-end;overflow:hidden;}}
.lp-hero video{{position:absolute;inset:0;width:100%;height:100%;object-fit:cover;opacity:.92;}}
.lp-hero-shade{{position:absolute;inset:0;background:linear-gradient(0deg,rgba(0,0,0,.85),rgba(0,0,0,.2) 55%,rgba(0,0,0,.05));}}
.lp-hero-in{{position:relative;max-width:1180px;margin:0 auto;width:100%;padding:0 28px 8%;text-shadow:0 1px 18px rgba(0,0,0,.5);}}
.lp-kick{{font-family:"Barlow Condensed",sans-serif;font-weight:600;letter-spacing:.3em;text-transform:uppercase;color:#E31E24;font-size:14px;}}
.lp-h1{{font-family:"Barlow Condensed",sans-serif;font-weight:600;text-transform:uppercase;letter-spacing:.04em;font-size:clamp(46px,8vw,96px);line-height:1;}}
.lp-hsub{{font-size:clamp(15px,2vw,19px);color:rgba(255,255,255,.85);max-width:560px;margin-top:8px;}}
.lp-scroll{{position:absolute;left:50%;bottom:18px;transform:translateX(-50%);color:rgba(255,255,255,.7);font-size:22px;animation:lpbob 1.8s ease-in-out infinite;}}
@keyframes lpbob{{0%,100%{{transform:translate(-50%,0)}}50%{{transform:translate(-50%,7px)}}}}
/* LP story */
.lp-story{{padding:30px 0 10px;}}
.lp-block{{display:grid;grid-template-columns:1.05fr .95fr;gap:46px;align-items:center;padding:46px 0;border-top:1px solid rgba(0,0,0,.08);}}
.lp-block:first-child{{border-top:0;}}
.lp-block-img{{border-radius:16px;overflow:hidden;background:#F0EEE9;aspect-ratio:4/3;}}
.lp-block-img img{{width:100%;height:100%;object-fit:cover;}}
.lp-block-kick{{font-family:"Barlow Condensed",sans-serif;font-weight:600;letter-spacing:.26em;text-transform:uppercase;color:#E31E24;font-size:13px;}}
.lp-block-h{{font-size:clamp(24px,3.2vw,34px);font-weight:700;line-height:1.4;margin-top:10px;}}
.lp-block-p{{font-size:15px;line-height:2;color:#444;margin-top:14px;}}
.lp-rev .lp-block-img{{order:2;}}
/* LP reels */
.lp-reels{{background:#0E0E0E;padding:54px 0;margin-top:20px;}}
.lp-reel-row{{display:flex;gap:14px;overflow-x:auto;margin-top:22px;scroll-snap-type:x mandatory;}}
.lp-reel{{flex:0 0 230px;aspect-ratio:9/16;border-radius:12px;overflow:hidden;background:#1a1a1a;scroll-snap-align:start;}}
.lp-reel video{{width:100%;height:100%;object-fit:cover;}}
/* LP support */
.lp-support{{display:flex;flex-wrap:wrap;gap:20px;align-items:center;justify-content:space-between;background:var(--mist,#F4F3F1);border-radius:16px;padding:26px 30px;}}
.lp-support-act{{display:flex;gap:12px;flex-wrap:wrap;}}
@media(max-width:760px){{
  .lp-hero{{aspect-ratio:4/5;}}
  .lp-block,.lp-block.lp-rev{{grid-template-columns:1fr;gap:18px;}}
  .lp-rev .lp-block-img{{order:0;}}
}}
</style>
</head>
<body class="font-sans text-[15px] leading-relaxed text-neutral-900 bg-white antialiased">
{NAV}
{lp_hero}
<div class="max-w-site mx-auto px-7 pt-6">
  <a href="products.html" class="inline-flex items-center gap-2 text-[13px] text-neutral-500 hover:text-shad transition"><i class="ti ti-arrow-left"></i>製品一覧</a>
</div>

<main class="max-w-site mx-auto px-7 py-8 grid md:grid-cols-2 gap-10 items-start">
  <div>
    <div class="g-main"><img id="gMain" src="{p['gallery'][0]}" alt="{code} {p['jp']}"></div>
    <div class="flex gap-3 mt-3">{thumbs}</div>
  </div>
  <div>
    <p class="flex items-center gap-2">{badges}<span class="font-disp text-[13px] tracking-[.22em] uppercase text-neutral-400">{p['series']}</span></p>
    <div class="flex items-end justify-between gap-4 mt-2">
      <h1 class="font-disp font-semibold text-[52px] leading-none tracking-[.04em] uppercase">{code}</h1>
      {f'<div class="text-right shrink-0">{cap_html}</div>' if cap_html else ''}
    </div>
    <p class="text-[15px] text-neutral-500 mt-1.5">{p['jp']}</p>
    <p class="text-[17px] font-bold mt-5 leading-relaxed">{html.escape(p['copy'])}</p>
    {feat_html}
    {f'<div class="flex flex-wrap gap-2 mt-5">{colors}</div>' if colors else ''}
    <div class="bg-ink rounded-[14px] p-5 mt-7 text-white">
      <p class="text-[13.5px] text-white/75 leading-relaxed">装着には車種専用フィッティングが必要です。適合確認済みの組み合わせをご案内します。</p>
      <div class="flex flex-wrap gap-3 mt-4">
        <a href="index.html#finder" class="btn bg-shad text-white hover:bg-[#c4151b] !py-3 !px-6 !text-[15px]"><i class="ti ti-motorbike"></i>適合を確認する</a>
        <a href="index.html#store" class="btn border border-white/40 text-white hover:border-white !py-3 !px-6 !text-[15px]">お近くの店舗を探す</a>
      </div>
    </div>
  </div>
</main>

{lp_story}
{lp_reels}

<section class="max-w-site mx-auto px-7 py-10 grid md:grid-cols-2 gap-10">
  <div>
    <h2 class="sec-ttl sec-ttl-quiet">Description</h2>
    <p class="text-[14.5px] leading-[2] mt-5">{desc}</p>
    {f'<p class="text-[13.5px] leading-[1.9] text-neutral-500 mt-4">{desc2}</p>' if desc2 else ''}
    {f'<p class="text-[12px] leading-[1.8] text-neutral-400 mt-5 border-t border-black/10 pt-4">{note}</p>' if note else ''}
  </div>
  <div>
    <h2 class="sec-ttl sec-ttl-quiet">Spec</h2>
    <table class="w-full mt-5 text-[14px]">{spec_rows(p['spec'])}</table>
  </div>
</section>

{lp_support}

{f"""<section class="bg-mist py-12 mt-8">
  <div class="max-w-site mx-auto px-7">
    <h2 class="sec-ttl sec-ttl-quiet">Same Series</h2>
    <div class="grid grid-cols-3 gap-5 mt-6">{rel_html}</div>
  </div>
</section>""" if rel_html else ''}

{FOOTER}
<script>
document.querySelectorAll('.g-thumb').forEach(function(b){{
  b.addEventListener('click',function(){{
    document.getElementById('gMain').src=b.dataset.src;
    document.querySelectorAll('.g-thumb').forEach(function(x){{x.classList.remove('on');}});
    b.classList.add('on');
  }});
}});
// ヒーロー動画は確実に再生
var hv=document.querySelector('.lp-hero video'); if(hv){{hv.play().catch(function(){{}});}}
// 縦リールはビューポート内のみ再生（preload none → 省データ）
var rv=document.querySelectorAll('.lp-reel video');
if(rv.length){{
  if('IntersectionObserver' in window){{
    var io=new IntersectionObserver(function(es){{es.forEach(function(e){{
      if(e.isIntersecting){{e.target.preload='auto';e.target.play().catch(function(){{}});}}else{{e.target.pause();}}
    }});}},{{threshold:.3}});
    rv.forEach(function(v){{io.observe(v);}});
  }} else {{ rv.forEach(function(v){{v.play().catch(function(){{}});}}); }}
}}
</script>
</body>
</html>'''
    open(f'{SITE}/product-{code.lower()}.html','w',encoding='utf-8').write(page)
    print('PAGE product-%s.html' % code.lower())

print('\nDONE:', len(products), 'pages')
